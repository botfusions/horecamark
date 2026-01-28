"""
Excel Report Generator for HorecaMark.

Generates daily Excel reports with multiple sheets:
- Summary (Ozet)
- Price Changes (Fiyat Degisiklikleri)
- Stock Changes (Stok Degisiklikleri)
- Price Comparison (Fiyat Karsilastirma)
- New Products (Yeni Urunler)
"""

from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, func, select
from sqlalchemy.orm import Session

from scraper.database import PriceChange, PriceSnapshot, Product, StockChange, get_session
from scraper.utils.config import Config
from scraper.utils.logger import get_logger

logger = get_logger("reporter")


# Site name mapping for Turkish headers
SITE_NAMES_TR = {
    "cafemarkt": "CafeMarkt",
    "arigastro": "AriGastro",
    "horecamarkt": "HorecaMarkt",
    "kariyermutfak": "KariyerMutfak",
    "mutbex": "Mutbex",
    "horecamark": "HorecaMark (Bizim)",
}

# Sheet names
SHEET_SUMMARY = "Ozet"
SHEET_PRICE_CHANGES = "Fiyat Degisiklikleri"
SHEET_STOCK_CHANGES = "Stok Degisiklikleri"
SHEET_PRICE_COMPARISON = "Fiyat Karsilastirma"
SHEET_NEW_PRODUCTS = "Yeni Urunler"

# Styling constants
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
BORDER_THIN = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_LEFT = Alignment(horizontal='left', vertical='center')
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')

# Conditional formatting colors
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")


@dataclass
class ReportSummary:
    """Summary statistics for the report."""

    date: date
    total_products: int
    price_changes: int
    price_decreases: int
    price_increases: int
    stock_changes: int
    new_products: int
    action_required: int

    def to_dict(self) -> dict:
        """Convert to dictionary."""
        return {
            "Tarih": self.date.strftime("%d.%m.%Y"),
            "Toplam Urun Tarandi": self.total_products,
            "Fiyat Degisikligi": self.price_changes,
            "Fiyat Dustu": self.price_decreases,
            "Fiyat Artti": self.price_increases,
            "Stok Degisikligi": self.stock_changes,
            "Yeni Urun": self.new_products,
            "Aksiyon Gerektiren": self.action_required,
        }


class ExcelReporter:
    """Excel report generator for HorecaMark."""

    def __init__(self, reports_dir: Optional[Path] = None):
        """Initialize reporter.

        Args:
            reports_dir: Directory to save reports (default: from Config)
        """
        self.reports_dir = reports_dir or Config.REPORTS_DIR
        self.reports_dir.mkdir(parents=True, exist_ok=True)

    def generate_daily_report(
        self,
        report_date: Optional[date] = None,
        session: Optional[Session] = None,
    ) -> Path:
        """Generate complete daily Excel report.

        Args:
            report_date: Date to generate report for (default: today)
            session: SQLAlchemy session (creates new if None)

        Returns:
            Path to generated Excel file
        """
        if report_date is None:
            report_date = date.today()

        close_session = False
        if session is None:
            session = get_session()
            close_session = True

        try:
            # Generate filename
            filename = f"daily_report_{report_date.strftime('%Y%m%d')}.xlsx"
            filepath = self.reports_dir / filename

            # Create workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Generate summary
            summary = self._generate_summary(session, report_date)

            # Create all sheets
            self._create_summary_sheet(wb, summary)
            self._create_price_changes_sheet(wb, session, report_date)
            self._create_stock_changes_sheet(wb, session, report_date)
            self._create_price_comparison_sheet(wb, session)
            self._create_new_products_sheet(wb, session, report_date)

            # Save workbook
            wb.save(filepath)
            logger.info(f"Report generated: {filepath}")

            return filepath

        finally:
            if close_session:
                session.close()

    def _generate_summary(self, session: Session, report_date: date) -> ReportSummary:
        """Generate report summary statistics.

        Args:
            session: SQLAlchemy session
            report_date: Date to generate summary for

        Returns:
            ReportSummary with statistics
        """
        day_start = datetime.combine(report_date, datetime.min.time())
        day_end = day_start + timedelta(days=1)

        # Count total products scraped
        total_products = (
            session.execute(
                select(func.count(func.distinct(PriceSnapshot.product_id)))
                .where(
                    and_(
                        PriceSnapshot.scraped_at >= day_start,
                        PriceSnapshot.scraped_at < day_end,
                    )
                )
            ).scalar()
            or 0
        )

        # Count price changes
        price_decreases = (
            session.execute(
                select(func.count(PriceChange.id))
                .where(
                    and_(
                        PriceChange.detected_at >= day_start,
                        PriceChange.detected_at < day_end,
                        PriceChange.change_percent < 0,
                    )
                )
            ).scalar()
            or 0
        )

        price_increases = (
            session.execute(
                select(func.count(PriceChange.id))
                .where(
                    and_(
                        PriceChange.detected_at >= day_start,
                        PriceChange.detected_at < day_end,
                        PriceChange.change_percent > 0,
                    )
                )
            ).scalar()
            or 0
        )

        price_changes = price_decreases + price_increases

        # Count stock changes
        stock_changes = (
            session.execute(
                select(func.count(StockChange.id))
                .where(
                    and_(
                        StockChange.detected_at >= day_start,
                        StockChange.detected_at < day_end,
                    )
                )
            ).scalar()
            or 0
        )

        # Count new products (first snapshot ever)
        new_products = (
            session.execute(
                select(func.count(func.distinct(PriceSnapshot.product_id)))
                .where(
                    and_(
                        PriceSnapshot.scraped_at >= day_start,
                        PriceSnapshot.scraped_at < day_end,
                    )
                )
                .having(func.count(PriceSnapshot.id) == 1)
            ).scalar()
            or 0
        )

        # Count critical actions (price decreases > 10%)
        action_required = (
            session.execute(
                select(func.count(PriceChange.id))
                .where(
                    and_(
                        PriceChange.detected_at >= day_start,
                        PriceChange.detected_at < day_end,
                        PriceChange.change_percent < -10,
                    )
                )
            ).scalar()
            or 0
        )

        return ReportSummary(
            date=report_date,
            total_products=total_products,
            price_changes=price_changes,
            price_decreases=price_decreases,
            price_increases=price_increases,
            stock_changes=stock_changes,
            new_products=new_products,
            action_required=action_required,
        )

    def _create_summary_sheet(self, wb: Workbook, summary: ReportSummary) -> None:
        """Create summary sheet.

        Args:
            wb: Workbook to add sheet to
            summary: ReportSummary data
        """
        ws = wb.create_sheet(SHEET_SUMMARY)

        # Title
        ws['B2'] = "HORECAMARK GUNLUK FIYAT ISTIHBARAT RAPORU"
        ws['B2'].font = Font(bold=True, size=14, color="4472C4")
        ws.merge_cells('B2:D2')

        # Date
        ws['B4'] = "Tarih:"
        ws['C4'] = summary.date.strftime("%d.%m.%Y")

        # Statistics
        row = 6
        data = summary.to_dict()

        for key, value in data.items():
            if key == "Tarih":
                continue
            ws[f'B{row}'] = key
            ws[f'B{row}'].font = Font(bold=True)
            ws[f'C{row}'] = value
            ws[f'C{row}'].number_format = '#,##0'
            row += 1

        # Formatting
        for col in range(2, 5):
            ws.column_dimensions[get_column_letter(col)].width = 20

        # Add borders
        for r in range(6, row):
            for c in range(2, 4):
                ws.cell(row=r, column=c).border = BORDER_THIN
                ws.cell(row=r, column=c).alignment = ALIGN_LEFT

    def _create_price_changes_sheet(
        self, wb: Workbook, session: Session, report_date: date
    ) -> None:
        """Create price changes sheet.

        Args:
            wb: Workbook to add sheet to
            session: SQLAlchemy session
            report_date: Report date
        """
        ws = wb.create_sheet(SHEET_PRICE_CHANGES)

        # Headers
        headers = [
            "Urun Adi",
            "Site",
            "Eski Fiyat",
            "Yeni Fiyat",
            "Degisim %",
            "Aksiyon Onerisi",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT_WHITE
            cell.fill = HEADER_FILL
            cell.border = BORDER_THIN
            cell.alignment = ALIGN_CENTER

        # Get price changes
        day_start = datetime.combine(report_date, datetime.min.time())
        day_end = day_start + timedelta(days=1)

        stmt = (
            select(PriceChange, Product)
            .join(Product, PriceChange.product_id == Product.id)
            .where(
                and_(
                    PriceChange.detected_at >= day_start,
                    PriceChange.detected_at < day_end,
                )
            )
            .order_by(PriceChange.change_percent.asc())
        )

        results = session.execute(stmt).all()

        # Fill data
        row = 2
        for change, product in results:
            change_pct = float(change.change_percent)

            ws.cell(row=row, column=1, value=product.normalized_name)
            ws.cell(row=row, column=2, value=SITE_NAMES_TR.get(change.site_name, change.site_name))
            ws.cell(row=row, column=3, value=float(change.old_price))
            ws.cell(row=row, column=4, value=float(change.new_price))
            ws.cell(row=row, column=5, value=change_pct)
            ws.cell(row=row, column=6, value=self._get_action_message(change_pct))

            # Conditional formatting
            fill = FILL_RED if change_pct < 0 else FILL_GREEN if change_pct > 5 else None
            if fill:
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = fill

            # Borders
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = BORDER_THIN

            row += 1

        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 35

        # Number formats
        for r in range(2, row):
            ws.cell(row=r, column=3).number_format = '#,##0.00 "TL"'
            ws.cell(row=r, column=4).number_format = '#,##0.00 "TL"'
            ws.cell(row=r, column=5).number_format = '0.00"%";(0.00)"%"'

    def _create_stock_changes_sheet(
        self, wb: Workbook, session: Session, report_date: date
    ) -> None:
        """Create stock changes sheet.

        Args:
            wb: Workbook to add sheet to
            session: SQLAlchemy session
            report_date: Report date
        """
        ws = wb.create_sheet(SHEET_STOCK_CHANGES)

        # Headers
        headers = [
            "Urun Adi",
            "Site",
            "Eski Durum",
            "Yeni Durum",
            "Degisiklik Turu",
            "Mesaj",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT_WHITE
            cell.fill = HEADER_FILL
            cell.border = BORDER_THIN
            cell.alignment = ALIGN_CENTER

        # Get stock changes
        day_start = datetime.combine(report_date, datetime.min.time())
        day_end = day_start + timedelta(days=1)

        stmt = (
            select(StockChange, Product)
            .join(Product, StockChange.product_id == Product.id)
            .where(
                and_(
                    StockChange.detected_at >= day_start,
                    StockChange.detected_at < day_end,
                )
            )
            .order_by(StockChange.detected_at.desc())
        )

        results = session.execute(stmt).all()

        # Fill data
        row = 2
        for change, product in results:
            ws.cell(row=row, column=1, value=product.normalized_name)
            ws.cell(row=row, column=2, value=SITE_NAMES_TR.get(change.site_name, change.site_name))
            ws.cell(row=row, column=3, value=change.previous_status or "-")
            ws.cell(row=row, column=4, value=change.new_status)
            ws.cell(row=row, column=5, value=self._translate_change_type(change.change_type))
            ws.cell(row=row, column=6, value=self._get_stock_message(change.change_type))

            # Conditional formatting based on change type
            fill = self._get_stock_fill(change.change_type)
            if fill:
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = fill

            # Borders
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = BORDER_THIN

            row += 1

        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 30

    def _create_price_comparison_sheet(self, wb: Workbook, session: Session) -> None:
        """Create price comparison pivot sheet.

        Args:
            wb: Workbook to add sheet to
            session: SQLAlchemy session
        """
        ws = wb.create_sheet(SHEET_PRICE_COMPARISON)

        # Headers
        headers = [
            "Urun Adi",
            "Marka",
            "Kategori",
            "CafeMarkt",
            "AriGastro",
            "HorecaMarkt",
            "KariyerMutfak",
            "Mutbex",
            "HorecaMark (Bizim)",
            "En Dusuk",
            "Fark %",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT_WHITE
            cell.fill = HEADER_FILL
            cell.border = BORDER_THIN
            cell.alignment = ALIGN_CENTER

        # Get products with recent prices
        cutoff = datetime.utcnow() - timedelta(days=7)

        stmt = (
            select(Product)
            .where(
                Product.id.in_(
                    select(PriceSnapshot.product_id).where(
                        PriceSnapshot.scraped_at >= cutoff
                    )
                )
            )
            .order_by(Product.normalized_name)
            .limit(1000)
        )

        products = session.execute(stmt).scalars().all()

        # Fill data
        row = 2
        site_keys = ["cafemarkt", "arigastro", "horecamarkt", "kariyermutfak", "mutbex", "horecamark"]

        for product in products:
            ws.cell(row=row, column=1, value=product.normalized_name)
            ws.cell(row=row, column=2, value=product.brand or "-")
            ws.cell(row=row, column=3, value=product.category or "-")

            prices = {}
            min_price = None
            our_price = None

            for col_idx, site_key in enumerate(site_keys, start=4):
                stmt = (
                    select(PriceSnapshot)
                    .where(
                        and_(
                            PriceSnapshot.product_id == product.id,
                            PriceSnapshot.site_name == site_key,
                            PriceSnapshot.scraped_at >= cutoff,
                        )
                    )
                    .order_by(PriceSnapshot.scraped_at.desc())
                    .limit(1)
                )

                snapshot = session.execute(stmt).scalar_one_or_none()

                if snapshot:
                    price_val = float(snapshot.price)
                    prices[site_key] = price_val
                    ws.cell(row=row, column=col_idx, value=price_val)

                    if site_key == "horecamark":
                        our_price = price_val

                    if min_price is None or price_val < min_price:
                        min_price = price_val
                else:
                    prices[site_key] = None
                    ws.cell(row=row, column=col_idx, value="-")

            # Find lowest price and mark it
            if min_price and our_price:
                min_col = 4
                for site_key in site_keys:
                    if prices.get(site_key) == min_price:
                        ws.cell(row=row, column=min_col).fill = FILL_GREEN
                        ws.cell(row=row, column=min_col).font = Font(bold=True, color="006100")
                    min_col += 1

                # Our price comparison
                ws.cell(row=row, column=10, value=min_price)

                diff_pct = ((our_price - min_price) / min_price) * 100 if min_price > 0 else 0
                ws.cell(row=row, column=11, value=diff_pct)

                if diff_pct > 10:
                    ws.cell(row=row, column=11).fill = FILL_RED
                elif diff_pct > 5:
                    ws.cell(row=row, column=11).fill = FILL_YELLOW

            # Borders and formats
            for col in range(1, 12):
                ws.cell(row=row, column=col).border = BORDER_THIN

            row += 1

        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        for col in range(4, 12):
            ws.column_dimensions[get_column_letter(col)].width = 12

        # Number formats
        for r in range(2, row):
            for col in range(4, 12):
                cell = ws.cell(row=r, column=col)
                if isinstance(cell.value, (int, float)):
                    if col == 11:
                        cell.number_format = '0.00"%";(0.00)"%"'
                    else:
                        cell.number_format = '#,##0.00 "TL"'

    def _create_new_products_sheet(
        self, wb: Workbook, session: Session, report_date: date
    ) -> None:
        """Create new products sheet.

        Args:
            wb: Workbook to add sheet to
            session: SQLAlchemy session
            report_date: Report date
        """
        ws = wb.create_sheet(SHEET_NEW_PRODUCTS)

        # Headers
        headers = ["Urun Adi", "Site", "Fiyat", "Stok Durumu", "URL"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT_WHITE
            cell.fill = HEADER_FILL
            cell.border = BORDER_THIN
            cell.alignment = ALIGN_CENTER

        # Get new products (first snapshot ever)
        day_start = datetime.combine(report_date, datetime.min.time())
        day_end = day_start + timedelta(days=1)

        # Find products with only one snapshot (newly discovered)
        subquery = (
            select(PriceSnapshot.product_id)
            .group_by(PriceSnapshot.product_id)
            .having(func.count(PriceSnapshot.id) == 1)
        )

        stmt = (
            select(PriceSnapshot, Product)
            .join(Product, PriceSnapshot.product_id == Product.id)
            .where(
                and_(
                    PriceSnapshot.product_id.in_(subquery),
                    PriceSnapshot.scraped_at >= day_start,
                    PriceSnapshot.scraped_at < day_end,
                )
            )
            .order_by(PriceSnapshot.site_name, Product.normalized_name)
        )

        results = session.execute(stmt).all()

        # Fill data
        row = 2
        for snapshot, product in results:
            ws.cell(row=row, column=1, value=product.normalized_name)
            ws.cell(row=row, column=2, value=SITE_NAMES_TR.get(snapshot.site_name, snapshot.site_name))
            ws.cell(row=row, column=3, value=float(snapshot.price))
            ws.cell(row=row, column=4, value=snapshot.stock_status or "-")
            ws.cell(row=row, column=5, value=snapshot.url or "-")

            # Borders
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = BORDER_THIN

            row += 1

        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 40

        # Number format for price
        for r in range(2, row):
            ws.cell(row=r, column=3).number_format = '#,##0.00 "TL"'

    def _get_action_message(self, change_percent: float) -> str:
        """Get action suggestion message.

        Args:
            change_percent: Price change percentage

        Returns:
            Action message in Turkish
        """
        if change_percent < -10:
            return "[ACIL] Rakip fiyatti dustu! Sen de dustur veya farklilastir."
        if change_percent < -5:
            return "[UYARI] Rakip hafif fiyat dustu. Izlemeye devam et."
        if change_percent > 10:
            return "[BILGI] Rakip fiyat artirdi. Marji koru, firsati degerlendir."
        if change_percent > 5:
            return "[NOT] Rakip hafif fiyat artirdi. Marji takip et."
        return "-"

    def _translate_change_type(self, change_type: str) -> str:
        """Translate change type to Turkish.

        Args:
            change_type: English change type

        Returns:
            Turkish translation
        """
        translations = {
            "stock_out": "Stok Tukendi",
            "stock_in": "Stok Geldi",
            "stock_low": "Stok Azaldi",
            "status_change": "Durum Degisikligi",
        }
        return translations.get(change_type, change_type)

    def _get_stock_message(self, change_type: str) -> str:
        """Get stock change message in Turkish.

        Args:
            change_type: Type of stock change

        Returns:
            Message in Turkish
        """
        messages = {
            "stock_out": "[FIRSAT] Rakip stoku tukendi! Satis firsati.",
            "stock_in": "[DIKKAT] Rakip stoku geldi. Rekabet basladi.",
            "stock_low": "[BILGI] Rakip stogu azaldi.",
            "status_change": "[BILGI] Stok durumu degisti.",
        }
        return messages.get(change_type, "-")

    def _get_stock_fill(self, change_type: str) -> Optional[PatternFill]:
        """Get fill color for stock change.

        Args:
            change_type: Type of stock change

        Returns:
            PatternFill or None
        """
        if change_type == "stock_out":
            return FILL_GREEN  # Opportunity
        if change_type == "stock_in":
            return FILL_YELLOW  # Warning
        if change_type == "stock_low":
            return FILL_YELLOW
        return None

    def cleanup_old_reports(self, keep_days: int = 30) -> list[Path]:
        """Remove reports older than keep_days.

        Args:
            keep_days: Number of days to keep reports

        Returns:
            List of removed file paths
        """
        cutoff = datetime.now() - timedelta(days=keep_days)
        removed = []

        for filepath in self.reports_dir.glob("daily_report_*.xlsx"):
            try:
                # Extract date from filename
                date_str = filepath.stem.replace("daily_report_", "")
                file_date = datetime.strptime(date_str, "%Y%m%d")

                if file_date < cutoff:
                    filepath.unlink()
                    removed.append(filepath)
                    logger.info(f"Removed old report: {filepath}")
            except ValueError:
                continue

        return removed


def generate_report(report_date: Optional[date] = None) -> Path:
    """Convenience function to generate daily report.

    Args:
        report_date: Date for report (default: today)

    Returns:
        Path to generated report file
    """
    reporter = ExcelReporter()
    return reporter.generate_daily_report(report_date)


if __name__ == "__main__":
    # Generate report for today
    report_path = generate_report()
    print(f"Report generated: {report_path}")
